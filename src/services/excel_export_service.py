from io import BytesIO
from typing import List, Optional

import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from src.core.logging import logger
from src.models.address import Coordinates
from src.models.market_analysis import MarketAnalysis
from src.services.google_location_service import GoogleLocationService


class ExcelStyles:
    LIGHT_BLUE = "DBE9F7"
    DARK_BLUE = "0070C0"
    BLUE = "0000FF"
    WHITE = "FFFFFF"
    DEFAULT_FONT = "Aptos Narrow"


class ExcelExportService:

    def __init__(self, location_service: Optional[GoogleLocationService] = None, satellite_zoom_level: int = 15):
        self._location_service = location_service
        self._satellite_zoom_level = satellite_zoom_level

    def export_market_analysis(self, market_analysis: MarketAnalysis) -> bytes:
        logger.info(f"Building Excel workbook from scratch for {market_analysis.address}")

        wb = openpyxl.Workbook()
        ws = wb.active

        address_parts = market_analysis.address.split(",")
        if len(address_parts) >= 3:
            city = address_parts[1].strip()
            state_zip = address_parts[2].strip()
            state = state_zip.split()[0] if state_zip else ""
            sheet_name = f"{city}, {state}"
        elif len(address_parts) >= 2:
            sheet_name = address_parts[1].strip()
        else:
            sheet_name = market_analysis.address[:31]

        ws.title = sheet_name
        ws.sheet_view.showGridLines = False

        self._write_header_section(ws, market_analysis)
        self._write_car_parc_section(ws, market_analysis)
        self._write_budget_section(ws, market_analysis)
        self._write_site_score_section(ws, market_analysis)
        self._write_market_summary_section(ws, market_analysis)
        self._write_key_stats_section(ws, market_analysis)
        self._write_competitors_section(ws, market_analysis)
        self._write_retail_performance_section(ws, market_analysis)
        self._write_warnings_section(ws, market_analysis)
        self._write_images_section(ws, market_analysis)

        ws.column_dimensions["A"].width = 8.63
        ws.column_dimensions["B"].width = 17.88
        ws.column_dimensions["C"].width = 9.25
        ws.column_dimensions["D"].width = 10.75
        ws.column_dimensions["E"].width = 13.0
        ws.column_dimensions["F"].width = 13.0
        ws.column_dimensions["G"].width = 13.0

        # Apply default font to all cells (preserving existing bold/color settings)
        self._apply_default_font(ws)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        logger.info("Excel workbook built successfully")
        return output.getvalue()

    def _write_header_section(self, ws, market_analysis: MarketAnalysis) -> None:
        ws["A2"] = "x"
        ws["B2"] = "Address"
        ws["C2"] = self._get_short_address(market_analysis.address)
        ws["D2"] = market_analysis.address

        self._apply_bold(ws, ["B2", "C2"])

    def _write_car_parc_section(self, ws, market_analysis: MarketAnalysis) -> None:
        ws["A4"] = "x"
        ws["B4"] = " "
        ws["E4"] = "Drive Time"
        ws["E4"].alignment = Alignment(horizontal="center")
        self._apply_bold(ws, ["E4"])

        # Sort all results by drive time
        all_results_sorted = sorted(market_analysis.car_parc_results, key=lambda x: x.drive_time_minutes)

        # Create hidden lookup table starting at column P (16)
        # Headers in row 4
        lookup_start_col = 16  # Column P
        ws.cell(row=4, column=lookup_start_col, value="Drive Time")
        ws.cell(row=4, column=lookup_start_col + 1, value="Car Parc")

        # Populate lookup table data starting at row 5
        lookup_data_start_row = 5
        for i, result in enumerate(all_results_sorted):
            row = lookup_data_start_row + i
            ws.cell(row=row, column=lookup_start_col, value=result.drive_time_minutes)
            ws.cell(row=row, column=lookup_start_col + 1, value=result.car_parc)

        lookup_data_end_row = lookup_data_start_row + len(all_results_sorted) - 1

        # Hide lookup table columns (P, Q)
        for col_idx in range(lookup_start_col, lookup_start_col + 2):
            ws.column_dimensions[get_column_letter(col_idx)].hidden = True

        # Create dropdown list from drive times
        drive_time_values = ",".join([str(int(r.drive_time_minutes)) for r in all_results_sorted])

        # Set up the three selectable columns (D, E, F) with dropdowns
        priority_defaults = [10, 12, 15]
        tam_percentages = [0.30, 0.25, 0.20]
        for i, default_time in enumerate(priority_defaults):
            col = 4 + i  # D=4, E=5, F=6
            col_letter = get_column_letter(col)

            # Row 5: Drive time dropdown
            cell = ws.cell(row=5, column=col)
            cell.value = default_time
            cell.number_format = '0 "Mins"'
            cell.font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True)
            cell.alignment = Alignment(horizontal="right")

            # Add dropdown validation
            dv = DataValidation(type="list", formula1=f'"{drive_time_values}"', allow_blank=False)
            dv.error = "Please select a valid drive time"
            dv.errorTitle = "Invalid Drive Time"
            ws.add_data_validation(dv)
            dv.add(cell)

            # Row 6: Car Parc - VLOOKUP
            lookup_range = f"$P${lookup_data_start_row}:$Q${lookup_data_end_row}"
            ws[f"{col_letter}6"] = f"=VLOOKUP({col_letter}5,{lookup_range},2,FALSE)"
            ws[f"{col_letter}6"].number_format = "#,##0"

            # Row 8: TAM % - constant values (30%, 25%, 20% for 10, 12, 15 min drive times)
            ws[f"{col_letter}8"] = tam_percentages[i]
            ws[f"{col_letter}8"].number_format = "0%"

            # Row 9: Total Addressable Market - formula
            ws[f"{col_letter}9"] = f"={col_letter}6*{col_letter}8"
            ws[f"{col_letter}9"].number_format = "#,##0"

            # Row 11: Washville Site Level Monthlies - formula
            ws[f"{col_letter}11"] = f"={col_letter}9*{col_letter}10"
            ws[f"{col_letter}11"].number_format = "#,##0"

        # Row 10: Market Share % - calculated from F63 (Current Share of Target)
        ws["D10"] = "=1-ROUNDUP(F63,1)"
        ws["D10"].number_format = "0%"
        ws["E10"] = "=D10-10%"
        ws["E10"].number_format = "0%"
        ws["F10"] = "=E10-10%"
        ws["F10"].number_format = "0%"

        self._apply_top_border(ws, 5, 4, 6, "dotted")
        ws["B6"] = "Total Car Parc"
        self._apply_top_border(ws, 6, 2, 6, "thin")

        ws["B7"] = "    Memo: Washville Median"
        ws["B7"].font = Font(name=ExcelStyles.DEFAULT_FONT, italic=True, size=9)

        # Washville Median values for 10, 12, 15 min drive times (columns D, E, F)
        median_values = [29500, 60000, 96000]
        for i, value in enumerate(median_values):
            cell = ws.cell(row=7, column=4 + i, value=value)
            cell.number_format = "#,##0"
            cell.font = Font(name=ExcelStyles.DEFAULT_FONT, italic=True, size=9)

        ws["B8"] = "Total Addressable Members"
        ws["B9"] = "Total Addressable Market"
        ws["B10"] = "Washville Market Share"
        ws["B11"] = "Washville Site Level Monthlies"

        self._apply_row_fill(ws, 11, 2, 6, ExcelStyles.LIGHT_BLUE)
        self._apply_row_font(ws, 11, 2, 6, bold=True)
        self._apply_box_border(ws, 11, 2, 6, "dotted")

    def _write_budget_section(self, ws, market_analysis: MarketAnalysis) -> None:
        if market_analysis.land_cost is not None:
            land_cost = market_analysis.land_cost
        else:
            land_cost = None

        ws["B13"] = "Preliminary Budget"
        self._apply_bold(ws, ["B13"])

        ws["B14"] = "Land Cost"
        ws["B15"] = "Build Cost"
        ws["B16"] = "Equipment/Install/DRB Cost"
        ws["B17"] = "Total Cost"
        ws["B18"] = "Expected EBITDA (Year 3)"
        ws["B19"] = "   Unlevered Cash-on-Cash Return (Target > 25%)"

        ws["F14"] = land_cost
        ws["F14"].number_format = "#,##0"
        ws["F15"] = 3000000
        ws["F15"].number_format = "#,##0"
        ws["F16"] = 1300000
        ws["F16"].number_format = "#,##0"
        ws["F17"] = "=SUM(F14:F16)"
        ws["F17"].number_format = "#,##0"
        ws["F18"].number_format = "#,##0"
        ws["F19"] = "=F18/F17"
        ws["F19"].number_format = "0.0%"

        self._apply_row_fill(ws, 17, 2, 6, ExcelStyles.LIGHT_BLUE)
        self._apply_row_fill(ws, 18, 2, 6, ExcelStyles.LIGHT_BLUE)
        self._apply_row_fill(ws, 19, 2, 6, ExcelStyles.LIGHT_BLUE)

        self._apply_row_font(ws, 17, 2, 6, bold=True)
        self._apply_row_font(ws, 18, 2, 6, bold=True)
        self._apply_row_font(ws, 19, 2, 6, bold=True)

        self._apply_multi_row_box_border(ws, 17, 19, 2, 6, "dotted")

        # Reference data section I15:K21
        blue_font = Font(name=ExcelStyles.DEFAULT_FONT, color=ExcelStyles.BLUE)

        # Row 15 - formula and value
        ws["I15"] = "=J15+F16"
        ws["I15"].number_format = "#,##0"
        ws["J15"] = 3200000
        ws["J15"].number_format = "#,##0"
        ws["J15"].font = blue_font

        # Row 16 - formula and value
        ws["I16"] = "=J16+F16"
        ws["I16"].number_format = "#,##0"
        ws["J16"] = 3000000
        ws["J16"].number_format = "#,##0"
        ws["J16"].font = blue_font

        # Row 17 - value only
        ws["J17"] = 2700000
        ws["J17"].number_format = "#,##0"
        ws["J17"].font = blue_font

        # Row 19 - Grade A+
        ws["I19"] = "A+"
        ws["J19"] = 3300000
        ws["J19"].number_format = "#,##0"
        ws["J19"].font = blue_font

        # Row 20 - Grade A
        ws["I20"] = "A"
        ws["J20"] = 3000000
        ws["J20"].number_format = "#,##0"
        ws["J20"].font = blue_font

        # Row 21 - Grade B
        ws["I21"] = "B"
        ws["J21"] = 2700000
        ws["J21"].number_format = "#,##0"
        ws["J21"].font = blue_font

    def _write_site_score_section(self, ws, market_analysis: MarketAnalysis) -> None:
        ws["A22"] = "x"
        ws["B22"] = "Site Score"
        self._apply_bold(ws, ["B22"])

        ws["C23"] = "Score"
        ws["C23"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, italic=True)
        ws["D23"] = "Weighting"
        ws["D23"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, italic=True)
        ws["E23"] = "Statistic"
        ws["E23"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, italic=True)
        ws["G23"] = "AM / PM"
        ws["G23"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, italic=True)
        self._apply_alignment(ws, ["C23", "D23", "E23", "G23"], "right")

        car_counts_value = market_analysis.traffic_counts if market_analysis.traffic_counts is not None else (market_analysis.key_stats.car_counts or "")

        criteria = [
            ("Car Counts", 0.075, '=TEXT(F24,"#,###")&" ("&G24&")"', None),
            ("Visibility", 0.15, "", None),
            ("Layout", 0.05, "", '0.0 "Acres"'),
            ("Ease", 0.05, "", None),
            ("Vac Space", 0.05, "", None),
            ("Retail", 0.125, "", None),
            ("Quality Competition", 0.2, "", None),
            ("Population ", 0.225, "=F6", "#,##0"),
            ("Income", 0.075, market_analysis.key_stats.median_income, '"$"#,##0'),
        ]

        row = 24
        for name, weighting, statistic, stat_format in criteria:
            ws[f"B{row}"] = name
            ws[f"C{row}"].number_format = "0.0"
            ws[f"D{row}"] = weighting
            ws[f"D{row}"].number_format = "0.0%"
            ws[f"D{row}"].font = Font(name=ExcelStyles.DEFAULT_FONT, italic=True)

            if statistic:
                if isinstance(statistic, str) and statistic.startswith("="):
                    ws[f"E{row}"] = statistic
                else:
                    ws[f"E{row}"] = statistic
            if stat_format:
                ws[f"E{row}"].number_format = stat_format

            row += 1

        ws["F24"] = car_counts_value
        ws["F24"].number_format = "#,##0"
        ws["G24"] = "PM"

        self._apply_top_border(ws, 24, 2, 5, "hair")

        grade_lookup = [
            ("F", 0),
            ("D-", 6),
            ("D", 6.333),
            ("D+", 6.667),
            ("C-", 7),
            ("C", 7.333),
            ("C+", 7.667),
            ("B-", 8),
            ("B", 8.333),
            ("B+", 8.667),
            ("A-", 9),
            ("A", 9.333),
            ("A+", 9.667),
        ]
        for i, (grade, value) in enumerate(grade_lookup):
            ws.cell(row=24 + i, column=10, value=grade)
            ws.cell(row=24 + i, column=11, value=value)
            ws.cell(row=24 + i, column=11).number_format = "0.000"

        score_rule = ColorScaleRule(
            start_type="num",
            start_value=5,
            start_color="FF5D5D",
            end_type="num",
            end_value=10,
            end_color="00B050",
        )
        ws.conditional_formatting.add("C24:C32", score_rule)

        ws["B33"] = "Weighted Score"
        ws["C33"] = "=MROUND(SUMPRODUCT(C24:C32,D24:D32),0.1)"
        ws["C33"].number_format = "0.0"
        self._apply_bold(ws, ["B33", "C33"])

        ws["B34"] = "Grade"
        ws["C34"] = "=LOOKUP(C33,K24:K36,J24:J36)"
        self._apply_alignment(ws, ["C34"], "right")
        self._apply_bold(ws, ["B34", "C34"])

        self._apply_row_fill(ws, 33, 2, 3, ExcelStyles.LIGHT_BLUE)
        self._apply_row_fill(ws, 34, 2, 3, ExcelStyles.LIGHT_BLUE)

        self._apply_multi_row_box_border(ws, 33, 34, 2, 3, "dotted")

    def _write_market_summary_section(self, ws, market_analysis: MarketAnalysis) -> None:
        ws["A37"] = "x"
        ws["B37"] = "Total Addressable Market"
        ws["F37"] = "=F9"
        ws["F37"].number_format = "#,##0"
        self._apply_bold(ws, ["B37"])

        ws["B38"] = "  Assumed Monthly Members / Car Wash"
        ws["B38"].font = Font(name=ExcelStyles.DEFAULT_FONT, italic=True)
        ws["F38"] = 5000
        ws["F38"].number_format = "#,##0"
        ws["F38"].font = Font(name=ExcelStyles.DEFAULT_FONT, italic=True)

        ws["B39"] = "Implied Car Washes Allowed in Market"
        ws["F39"] = "=F37/F38"
        ws["F39"].number_format = "0.0"

        ws["B40"] = "Current Monthlies in Market"
        ws["F40"] = "=F62"
        ws["F40"].number_format = "#,##0"

        ws["B41"] = "Remaining Monthlies in Market"
        ws["F41"] = "=F37-F40"
        ws["F41"].number_format = "#,##0"

        ws["B42"] = "Implied Remaining Car Washes Allowed in Market"
        ws["F42"] = "=F41/F38"
        ws["F42"].number_format = "0.0"

        self._apply_row_fill(ws, 39, 2, 6, ExcelStyles.LIGHT_BLUE)
        self._apply_row_font(ws, 39, 2, 6, bold=True)

        self._apply_row_fill(ws, 42, 2, 6, ExcelStyles.DARK_BLUE)
        self._apply_row_font(ws, 42, 2, 6, color=ExcelStyles.WHITE, bold=True)

    def _write_key_stats_section(self, ws, market_analysis: MarketAnalysis) -> None:
        ws["D44"] = "Washville"
        ws["D44"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, size=9)
        self._apply_alignment(ws, ["D44"], "right")
        ws["A45"] = "x"
        ws["B45"] = "Key Stats"
        ws["C45"] = self._get_short_address(market_analysis.address)
        ws["C45"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, size=9)
        ws["D45"] = "Median"
        ws["D45"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, size=9)
        self._apply_bold(ws, ["B45"])
        self._apply_alignment(ws, ["C45", "D45"], "right")

        ws["B46"] = "Car Counts"
        ws["C46"] = "=F24"
        ws["C46"].number_format = "#,##0"
        ws["D46"] = 23000
        ws["D46"].number_format = "#,##0"
        ws["D46"].font = Font(name=ExcelStyles.DEFAULT_FONT, italic=True)

        self._apply_top_border(ws, 46, 2, 4, "dotted")

        ws["B47"] = "Car Parc"
        ws["C47"] = "=F6"
        ws["C47"].number_format = "#,##0"
        ws["D47"] = 96000
        ws["D47"].number_format = "#,##0"
        ws["D47"].font = Font(name=ExcelStyles.DEFAULT_FONT, italic=True)

        ws["B48"] = "Median Income"
        ws["C48"] = "=E32"
        ws["C48"].number_format = "#,##0"
        ws["D48"] = 85000
        ws["D48"].number_format = "#,##0"
        ws["D48"].font = Font(name=ExcelStyles.DEFAULT_FONT, italic=True)

        ws["B49"] = "Average Age"
        ws["C49"] = round(market_analysis.key_stats.median_age)
        ws["D49"] = 42
        ws["D49"].number_format = "#,##0"
        ws["D49"].font = Font(name=ExcelStyles.DEFAULT_FONT, italic=True)

        ws["B50"] = "Snowfall (in.)"
        ws["D50"] = 55.3
        ws["D50"].number_format = "0.0"
        ws["D50"].font = Font(name=ExcelStyles.DEFAULT_FONT, italic=True)

        self._apply_fill(ws, ["C46", "C47", "C48", "C49", "C50"], ExcelStyles.LIGHT_BLUE)
        self._apply_bold(ws, ["C46", "C47", "C48", "C49", "C50"])

        self._apply_bottom_border(ws, 50, 2, 4, "dotted")

    def _write_competitors_section(self, ws, market_analysis: MarketAnalysis) -> None:
        ws["C53"] = "70%"
        ws["D53"] = "Total"
        ws["E53"] = "%"
        ws["F53"] = "Members in"
        ws["G53"] = "%"
        self._apply_bold(ws, ["C53", "D53", "E53", "F53", "G53"])
        self._apply_alignment(ws, ["C53", "D53", "E53", "F53", "G53"], "center")

        ws["A54"] = "x"
        ws["B54"] = "Competitors"
        ws["C54"] = "Car Parc"
        ws["D54"] = "Members"
        ws["E54"] = "Overlap"
        ws["F54"] = "Market"
        ws["G54"] = "Market Share"
        self._apply_bold(ws, ["B54", "C54", "D54", "E54", "F54", "G54"])
        self._apply_alignment(ws, ["C54", "D54", "E54", "F54", "G54"], "center")

        sorted_competitors = sorted(market_analysis.competitors, key=lambda x: x.distance_miles)

        row = 55
        for competitor in sorted_competitors[:7]:
            ws[f"B{row}"] = competitor.name

            if competitor.car_parc:
                ws[f"C{row}"] = competitor.car_parc
                ws[f"C{row}"].number_format = "#,##0"

            ws[f"D{row}"] = competitor.total_members
            ws[f"D{row}"].number_format = "#,##0"

            ws[f"E{row}"] = competitor.overlap_percentage / 100
            ws[f"E{row}"].number_format = "0%"

            ws[f"F{row}"] = f"=D{row}*E{row}"
            ws[f"F{row}"].number_format = "#,##0"

            ws[f"G{row}"] = f"=F{row}/$F$62"
            ws[f"G{row}"].number_format = "0%"

            row += 1

        for r in range(row, 62):
            ws[f"F{r}"] = f"=D{r}*E{r}"
            ws[f"F{r}"].number_format = "#,##0"
            ws[f"G{r}"] = f"=F{r}/$F$62"
            ws[f"G{r}"].number_format = "0%"

        ws["B62"] = "Total Market Members"
        ws["F62"] = "=SUM(F55:F61)"
        ws["F62"].number_format = "#,##0"
        self._apply_bold(ws, ["B62"])

        ws["B63"] = "Current Share of Target (>50% = Market Share Battle)"
        ws["B63"].font = Font(name=ExcelStyles.DEFAULT_FONT, italic=True)
        ws["F63"] = "=F62/F9"
        ws["F63"].number_format = "0%"
        ws["F63"].font = Font(name=ExcelStyles.DEFAULT_FONT, italic=True)

    def _write_retail_performance_section(self, ws, market_analysis: MarketAnalysis) -> int:
        ws["A65"] = "x"
        ws["B65"] = "Retail Performance"
        self._apply_bold(ws, ["B65"])

        ws["C66"] = "National"
        ws["C66"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, italic=True, size=9)
        ws["D66"] = "State"
        ws["D66"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, italic=True, size=9)
        ws["E66"] = "Distance"
        ws["E66"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, italic=True, size=9)
        ws["C67"] = "Percentile"
        ws["C67"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, italic=True, size=9)
        ws["D67"] = "Percentile"
        ws["D67"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, italic=True, size=9)
        ws["E67"] = "(miles)"
        ws["E67"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, italic=True, size=9)
        ws["F67"] = "Visits"
        ws["F67"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, italic=True, size=9)

        self._apply_alignment(ws, ["C66", "D66", "E66", "C67", "D67", "E67", "F67"], "center")

        ws["B67"] = "Retailer"
        ws["B67"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, italic=True)

        row = 68
        data_start_row = row

        if market_analysis.reference_poi_retail:
            poi = market_analysis.reference_poi_retail
            ws[f"A{row}"] = "*"
            ws[f"A{row}"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True, color="FF0000")
            ws[f"B{row}"] = poi.name
            ws[f"B{row}"].font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True)

            if poi.national_percentile is not None:
                ws[f"C{row}"] = poi.national_percentile
                ws[f"C{row}"].number_format = "0%"

            if poi.state_percentile is not None:
                ws[f"D{row}"] = poi.state_percentile
                ws[f"D{row}"].number_format = "0%"

            ws[f"E{row}"] = 0
            ws[f"E{row}"].number_format = "0.0"

            if poi.visits is not None:
                ws[f"F{row}"] = poi.visits
                ws[f"F{row}"].number_format = "#,##0"

            row += 1

        for retailer in market_analysis.retailers:
            ws[f"B{row}"] = retailer.name

            if retailer.national_percentile is not None:
                ws[f"C{row}"] = retailer.national_percentile
                ws[f"C{row}"].number_format = "0%"

            if retailer.state_percentile is not None:
                ws[f"D{row}"] = retailer.state_percentile
                ws[f"D{row}"].number_format = "0%"

            ws[f"E{row}"] = retailer.distance_miles
            ws[f"E{row}"].number_format = "0.0"

            if retailer.visits is not None:
                ws[f"F{row}"] = retailer.visits
                ws[f"F{row}"].number_format = "#,##0"

            row += 1

        last_row = row - 1

        percentile_rule = ColorScaleRule(
            start_type="num",
            start_value=0,
            start_color="F8696B",
            mid_type="num",
            mid_value=0.5,
            mid_color="FFEB84",
            end_type="num",
            end_value=1,
            end_color="63BE7B",
        )
        ws.conditional_formatting.add(f"C{data_start_row}:D{last_row}", percentile_rule)

        if data_start_row <= last_row:
            distance_rule = ColorScaleRule(
                start_type="num",
                start_value=0,
                start_color="63BE7B",
                mid_type="num",
                mid_value=1.5,
                mid_color="FFEB84",
                end_type="num",
                end_value=3,
                end_color="F8696B",
            )
            ws.conditional_formatting.add(f"E{data_start_row}:E{last_row}", distance_rule)

        total_row = last_row + 1
        ws[f"B{total_row}"] = "Total"
        ws[f"F{total_row}"] = f"=SUM(F{data_start_row}:F{last_row})"
        ws[f"F{total_row}"].number_format = "#,##0"
        self._apply_bold(ws, [f"B{total_row}", f"F{total_row}"])
        self._apply_top_border(ws, total_row, 2, 6, "thin")

        retailer_count = len(market_analysis.retailers) + (1 if market_analysis.reference_poi_retail else 0)
        logger.debug(f"Retail performance section created with {retailer_count} retailers (including POI)")
        return total_row

    def _write_warnings_section(self, ws, market_analysis: MarketAnalysis) -> None:
        if not market_analysis.warnings:
            return

        logger.info(f"Writing {len(market_analysis.warnings)} warnings to Excel")

        # Find the last used row to place warnings below all data
        last_row = ws.max_row + 2

        ws.cell(row=last_row, column=2, value="Warnings")
        ws.cell(row=last_row, column=2).font = Font(
            name=ExcelStyles.DEFAULT_FONT, bold=True, color="C0392B"
        )

        warning_fill = PatternFill(start_color="FEF0EC", end_color="FEF0EC", fill_type="solid")
        warning_font = Font(name=ExcelStyles.DEFAULT_FONT, color="C0392B", size=9)

        for i, warning in enumerate(market_analysis.warnings):
            row = last_row + 1 + i
            cell = ws.cell(row=row, column=2, value=f"⚠ {warning}")
            cell.font = warning_font
            cell.fill = warning_fill
            # Merge across columns B-F so the warning text has room
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)

    def _write_images_section(self, ws, market_analysis: MarketAnalysis) -> None:
        if not self._location_service:
            logger.info("No location service provided, skipping image section")
            return

        coordinates = Coordinates(
            latitude=market_analysis.latitude,
            longitude=market_analysis.longitude,
            formatted_address=market_analysis.address,
        )

        street_view_bytes = self._location_service.download_street_view_image(coordinates)
        satellite_zoomed_in_bytes = self._location_service.download_satellite_image(coordinates, 17)
        satellite_zoomed_out_bytes = self._location_service.download_satellite_image(coordinates, 14)

        # Images on row 1, starting at column K
        # Order: furthest satellite, closer satellite, street view
        if satellite_zoomed_out_bytes:
            satellite_zoomed_out_img = Image(BytesIO(satellite_zoomed_out_bytes))
            satellite_zoomed_out_img.anchor = "K1"
            ws.add_image(satellite_zoomed_out_img)
            logger.info("Satellite zoomed-out image added to Excel at K1")
        else:
            logger.warning("Failed to download satellite zoomed-out image")

        if satellite_zoomed_in_bytes:
            satellite_zoomed_in_img = Image(BytesIO(satellite_zoomed_in_bytes))
            satellite_zoomed_in_img.anchor = "Y1"
            ws.add_image(satellite_zoomed_in_img)
            logger.info("Satellite zoomed-in image added to Excel at Y1")
        else:
            logger.warning("Failed to download satellite zoomed-in image")

        if street_view_bytes:
            street_view_img = Image(BytesIO(street_view_bytes))
            street_view_img.anchor = "AI1"
            ws.add_image(street_view_img)
            logger.info("Street view image added to Excel at AI1")
        else:
            logger.warning("Failed to download street view image")

    def _get_short_address(self, full_address: str) -> str:
        address_parts = full_address.split(",")
        if len(address_parts) >= 2:
            return address_parts[1].strip()
        return full_address

    def _apply_bold(self, ws, cells: List[str]) -> None:
        bold_font = Font(name=ExcelStyles.DEFAULT_FONT, bold=True)
        for cell in cells:
            ws[cell].font = bold_font

    def _apply_fill(self, ws, cells: List[str], color: str) -> None:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for cell in cells:
            ws[cell].fill = fill

    def _apply_row_fill(self, ws, row: int, start_col: int, end_col: int, color: str) -> None:
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).fill = fill

    def _apply_row_font(self, ws, row: int, start_col: int, end_col: int, color: str = None, bold: bool = False) -> None:
        font = Font(name=ExcelStyles.DEFAULT_FONT, color=color, bold=bold)
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).font = font

    def _apply_top_border(self, ws, row: int, start_col: int, end_col: int, style: str = "thin") -> None:
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(top=Side(style=style))

    def _apply_box_border(self, ws, row: int, start_col: int, end_col: int, style: str = "dotted") -> None:
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            borders = {}

            if col == start_col:
                borders["left"] = Side(style=style)
            if col == end_col:
                borders["right"] = Side(style=style)

            borders["top"] = Side(style=style)
            borders["bottom"] = Side(style=style)

            cell.border = Border(**borders)

    def _apply_multi_row_box_border(self, ws, start_row: int, end_row: int, start_col: int, end_col: int, style: str = "dotted") -> None:
        for row in range(start_row, end_row + 1):
            for col in range(start_col, end_col + 1):
                cell = ws.cell(row=row, column=col)
                borders = {}

                # Left border only on first column
                if col == start_col:
                    borders["left"] = Side(style=style)
                # Right border only on last column
                if col == end_col:
                    borders["right"] = Side(style=style)
                # Top border only on first row
                if row == start_row:
                    borders["top"] = Side(style=style)
                # Bottom border only on last row
                if row == end_row:
                    borders["bottom"] = Side(style=style)

                if borders:
                    cell.border = Border(**borders)

    def _apply_bottom_border(self, ws, row: int, start_col: int, end_col: int, style: str = "hair") -> None:
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(bottom=Side(style=style))

    def _apply_alignment(self, ws, cells: List[str], horizontal: str) -> None:
        alignment = Alignment(horizontal=horizontal)
        for cell in cells:
            ws[cell].alignment = alignment

    def _apply_default_font(self, ws) -> None:
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    current_font = cell.font
                    cell.font = Font(
                        name=ExcelStyles.DEFAULT_FONT,
                        bold=current_font.bold if current_font else False,
                        italic=current_font.italic if current_font else False,
                        color=current_font.color if current_font else None,
                        size=current_font.size if current_font and current_font.size else None,
                    )
